import openpyxl
from operator import itemgetter
import datetime

wb = openpyxl.load_workbook('C:/Users/Damian/Desktop/ranking.xlsx')
ws = wb['Arkusz1']
ws_wyniki = wb['Wyniki']
range_nicki = ws['A3':'A51']

lista_graczy = []
for cell in range_nicki:
    lista_graczy.append(cell[0].value)

typowanie = {}
for nick in lista_graczy:
    typowanie[nick] = {}

range_w_dol = range(3, 65)
for row in range_w_dol:
    nick = ws.cell(row=row, column=1).value
    col = 2
    for nr_meczu in range(1, 100):
        if not ws.cell(row=row, column=col).value:
            break
        typowanie[nick][nr_meczu] = ws.cell(row=row, column=col).value
        col += 1

wyniki = {}
col = 2
for nr_meczu in range(1, 100):
    if not ws_wyniki.cell(row=3, column=col).value:
        break
    wyniki[nr_meczu] = ws_wyniki.cell(row=3, column=col).value
    col += 1

ranking = {}
for gracz in typowanie:
    ranking[gracz] = 0
    for nr_meczu in range(1, len(wyniki) + 1):
        wynik = ''
        typ = ''
        if wyniki[nr_meczu][0] > wyniki[nr_meczu][2]:
            wynik = 'A'
        elif wyniki[nr_meczu][0] < wyniki[nr_meczu][2]:
            wynik = 'B'
        elif wyniki[nr_meczu][0] == wyniki[nr_meczu][2]:
            wynik = 'AB'
        if typowanie[gracz][nr_meczu][0] > typowanie[gracz][nr_meczu][2]:
            typ = 'A'
        elif typowanie[gracz][nr_meczu][0] < typowanie[gracz][nr_meczu][2]:
            typ = 'B'
        elif typowanie[gracz][nr_meczu][0] == typowanie[gracz][nr_meczu][2]:
            typ = 'AB'
        if typowanie[gracz][nr_meczu] == wyniki[nr_meczu]:
            ranking[gracz] += 3
        if wynik == typ:
            ranking[gracz] += 2

lista_koncowa = sorted(list(ranking.items()), key=itemgetter(1))
lista_koncowa.reverse()

miejsce = 1

data_aktualna = datetime.datetime.now()
print('Data aktualizacji: ', end='')
print(data_aktualna.strftime('%Y-%m-%d %H:%M:%S'))

for i in lista_koncowa:
    print(miejsce, end='. ')
    print(i)
    miejsce += 1
